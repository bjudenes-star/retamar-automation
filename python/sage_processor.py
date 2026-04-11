"""Procesador de ficheros Sage 50: CSV de facturas, SEPA XML y Excel manual.

Uso:
    python -m python.sage_processor --input-dir ./datos
    python -m python.sage_processor --watch ./datos
    python -m python.sage_processor --help
"""

import argparse
import csv
import hashlib
import logging
import os
import sqlite3
import sys
import time
import xml.etree.ElementTree as ET
from datetime import datetime
from pathlib import Path

try:
    import openpyxl
except ImportError:
    openpyxl = None

DB_PATH = Path(__file__).parent.parent / "db" / "retamar.db"

logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s [%(levelname)s] %(message)s",
    datefmt="%Y-%m-%d %H:%M:%S",
)
log = logging.getLogger("sage_processor")

# Mapeo flexible de columnas CSV de Sage (variantes conocidas)
COLUMN_ALIASES = {
    "numero_factura": ["Num.", "N\u00ba Factura", "Numero", "N\u00famero", "NumFactura"],
    "fecha": ["Fecha", "Fecha Factura", "FechaFactura"],
    "family_code": ["Cliente", "C\u00f3digo Cliente", "CodCliente", "Cod. Cliente"],
    "nombre": ["Nombre", "Nombre Cliente", "Raz\u00f3n Social"],
    "nif": ["NIF", "CIF", "DNI", "NIF/CIF"],
    "base_imponible": ["Base Imponible", "Base", "BaseImp"],
    "iva": ["IVA", "Cuota IVA", "IVA %"],
    "total": ["Total", "Total Factura", "Importe"],
    "fecha_vencimiento": ["Vencimiento", "Fecha Vencimiento", "FechaVto"],
    "estado": ["Estado", "Situaci\u00f3n"],
}


def _hash_contenido(data: str) -> str:
    """Genera SHA256 del contenido para detectar duplicados."""
    return hashlib.sha256(data.encode("utf-8")).hexdigest()


def init_db(db_path: Path = DB_PATH) -> sqlite3.Connection:
    """Inicializa la base de datos con el esquema completo."""
    conn = sqlite3.connect(str(db_path))
    conn.execute("PRAGMA journal_mode=WAL")
    schema_path = Path(__file__).parent.parent / "db" / "schema.sql"
    if schema_path.exists():
        conn.executescript(schema_path.read_text(encoding="utf-8"))
    return conn


def _resolve_columns(header: list[str]) -> dict[str, str]:
    """Mapea columnas del CSV a nombres internos usando alias conocidos."""
    mapping = {}
    header_stripped = [h.strip() for h in header]
    for internal_name, aliases in COLUMN_ALIASES.items():
        for alias in aliases:
            if alias in header_stripped:
                mapping[internal_name] = alias
                break
    return mapping


def _parse_fecha(valor: str) -> str | None:
    """Intenta parsear una fecha en formatos comunes espa\u00f1oles."""
    if not valor:
        return None
    for fmt in ("%d/%m/%Y", "%d-%m-%Y", "%Y-%m-%d", "%d/%m/%y"):
        try:
            return datetime.strptime(valor.strip(), fmt).strftime("%Y-%m-%d")
        except ValueError:
            continue
    return None


def _parse_importe(valor: str) -> float:
    """Parsea importes con formato espa\u00f1ol (1.234,56) o est\u00e1ndar."""
    if not valor:
        return 0.0
    valor = valor.strip().replace("\u20ac", "").replace(" ", "")
    if "," in valor and "." in valor:
        valor = valor.replace(".", "").replace(",", ".")
    elif "," in valor:
        valor = valor.replace(",", ".")
    return float(valor)


def _upsert_familia(conn: sqlite3.Connection, codigo: str, nombre: str,
                    nif: str | None = None):
    """Inserta o actualiza una familia en la base de datos."""
    conn.execute(
        """INSERT INTO familias (codigo_familia, nombre, dni)
           VALUES (?, ?, ?)
           ON CONFLICT(codigo_familia) DO UPDATE SET
               nombre = COALESCE(excluded.nombre, familias.nombre),
               dni = COALESCE(excluded.dni, familias.dni)""",
        (codigo, nombre, nif),
    )


def _insert_factura(conn: sqlite3.Connection, factura: dict) -> bool:
    """Inserta factura si no existe (por hash). Devuelve True si se insert\u00f3."""
    try:
        conn.execute(
            """INSERT INTO facturas
               (numero_factura, family_code, fecha, concepto, base_imponible,
                iva, total, fecha_vencimiento, estado, hash_contenido, origen)
               VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)""",
            (
                factura["numero_factura"],
                factura["family_code"],
                factura.get("fecha"),
                factura.get("concepto"),
                factura.get("base_imponible"),
                factura.get("iva"),
                factura["total"],
                factura.get("fecha_vencimiento"),
                factura.get("estado", "pendiente"),
                factura["hash_contenido"],
                factura.get("origen"),
            ),
        )
        return True
    except sqlite3.IntegrityError:
        return False


def procesar_csv(filepath: Path, conn: sqlite3.Connection) -> dict:
    """Procesa un CSV de facturas de Sage 50."""
    log.info("Procesando CSV: %s", filepath.name)
    stats = {"nuevas": 0, "duplicadas": 0, "errores": 0}

    with open(filepath, encoding="utf-8-sig", newline="") as f:
        reader = csv.DictReader(f, delimiter=";")
        col_map = _resolve_columns(reader.fieldnames or [])

        if "family_code" not in col_map or "total" not in col_map:
            log.error("CSV sin columnas obligatorias (cliente/total): %s", filepath.name)
            stats["errores"] = 1
            return stats

        for row in reader:
            try:
                codigo = row[col_map["family_code"]].strip()
                nombre = row.get(col_map.get("nombre", ""), "").strip()
                nif = row.get(col_map.get("nif", ""), "").strip() or None

                raw = ";".join(str(row[k]) for k in sorted(row.keys()))
                hash_c = _hash_contenido(raw)

                factura = {
                    "numero_factura": row.get(col_map.get("numero_factura", ""), hash_c[:12]).strip(),
                    "family_code": codigo,
                    "fecha": _parse_fecha(row.get(col_map.get("fecha", ""), "")),
                    "concepto": None,
                    "base_imponible": _parse_importe(row.get(col_map.get("base_imponible", ""), "0")),
                    "iva": _parse_importe(row.get(col_map.get("iva", ""), "0")),
                    "total": _parse_importe(row[col_map["total"]]),
                    "fecha_vencimiento": _parse_fecha(row.get(col_map.get("fecha_vencimiento", ""), "")),
                    "estado": row.get(col_map.get("estado", ""), "pendiente").strip().lower() or "pendiente",
                    "hash_contenido": hash_c,
                    "origen": "csv",
                }

                _upsert_familia(conn, codigo, nombre, nif)
                if _insert_factura(conn, factura):
                    stats["nuevas"] += 1
                else:
                    stats["duplicadas"] += 1
            except Exception as e:
                log.error("Error en fila CSV: %s", e)
                stats["errores"] += 1

    conn.commit()
    log.info("CSV %s: %d nuevas, %d duplicadas, %d errores",
             filepath.name, stats["nuevas"], stats["duplicadas"], stats["errores"])
    return stats


def procesar_sepa_xml(filepath: Path, conn: sqlite3.Connection) -> dict:
    """Procesa un fichero SEPA XML pain.008 (remesa de cobro)."""
    log.info("Procesando SEPA XML: %s", filepath.name)
    stats = {"nuevas": 0, "duplicadas": 0, "errores": 0}

    ns = {"sepa": "urn:iso:std:iso:20022:tech:xsd:pain.008.001.02"}
    tree = ET.parse(filepath)
    root = tree.getroot()

    # Detectar namespace din\u00e1micamente
    tag = root.tag
    if "{" in tag:
        ns_uri = tag.split("}")[0].strip("{")
        ns = {"sepa": ns_uri}

    for tx in root.iter(f"{{{ns['sepa']}}}DrctDbtTxInf"):
        try:
            # Referencia del recibo
            ref_el = tx.find(f".//{{{ns['sepa']}}}EndToEndId")
            ref = ref_el.text if ref_el is not None else None

            # Importe
            amt_el = tx.find(f".//{{{ns['sepa']}}}InstdAmt")
            importe = float(amt_el.text) if amt_el is not None else 0.0

            # Nombre del deudor
            nm_el = tx.find(f".//{{{ns['sepa']}}}Dbtr/{{{ns['sepa']}}}Nm")
            nombre = nm_el.text if nm_el is not None else ""

            # IBAN del deudor
            iban_el = tx.find(f".//{{{ns['sepa']}}}DbtrAcct/{{{ns['sepa']}}}Id/{{{ns['sepa']}}}IBAN")

            # C\u00f3digo familia: buscar en referencia o usar primeros 8 chars
            family_code = ""
            if ref:
                import re
                match = re.search(r"\d{8}", ref)
                if match:
                    family_code = match.group(0)

            if not family_code:
                log.warning("No se pudo extraer c\u00f3digo familia de referencia: %s", ref)
                stats["errores"] += 1
                continue

            raw = f"{ref}|{importe}|{nombre}|{family_code}"
            hash_c = _hash_contenido(raw)

            _upsert_familia(conn, family_code, nombre)

            factura = {
                "numero_factura": ref or hash_c[:12],
                "family_code": family_code,
                "fecha": datetime.now().strftime("%Y-%m-%d"),
                "concepto": "Remesa SEPA",
                "base_imponible": None,
                "iva": None,
                "total": importe,
                "fecha_vencimiento": None,
                "estado": "pendiente",
                "hash_contenido": hash_c,
                "origen": "sepa",
            }

            if _insert_factura(conn, factura):
                stats["nuevas"] += 1
            else:
                stats["duplicadas"] += 1
        except Exception as e:
            log.error("Error en transacci\u00f3n SEPA: %s", e)
            stats["errores"] += 1

    conn.commit()
    log.info("SEPA %s: %d nuevas, %d duplicadas, %d errores",
             filepath.name, stats["nuevas"], stats["duplicadas"], stats["errores"])
    return stats


def procesar_excel(filepath: Path, conn: sqlite3.Connection) -> dict:
    """Procesa un Excel manual (comedor, FP, extraescolares)."""
    if openpyxl is None:
        log.error("openpyxl no instalado. Ejecuta: pip install openpyxl")
        return {"nuevas": 0, "duplicadas": 0, "errores": 1}

    log.info("Procesando Excel: %s", filepath.name)
    stats = {"nuevas": 0, "duplicadas": 0, "errores": 0}

    wb = openpyxl.load_workbook(filepath, read_only=True, data_only=True)
    for sheet in wb.sheetnames:
        ws = wb[sheet]
        rows = list(ws.iter_rows(values_only=True))
        if not rows:
            continue

        header = [str(h or "").strip() for h in rows[0]]
        col_map = _resolve_columns(header)

        if "family_code" not in col_map or "total" not in col_map:
            log.warning("Hoja '%s' sin columnas obligatorias, saltando", sheet)
            continue

        for i, row in enumerate(rows[1:], start=2):
            try:
                data = dict(zip(header, row))
                codigo = str(data.get(col_map["family_code"], "")).strip()
                if not codigo:
                    continue

                nombre = str(data.get(col_map.get("nombre", ""), "")).strip()
                total_val = data.get(col_map["total"], 0)
                if isinstance(total_val, str):
                    total_val = _parse_importe(total_val)
                total_val = float(total_val or 0)

                raw = f"{filepath.name}|{sheet}|{i}|{codigo}|{total_val}"
                hash_c = _hash_contenido(raw)

                _upsert_familia(conn, codigo, nombre or "Sin nombre")

                fecha_raw = data.get(col_map.get("fecha", ""), "")
                if isinstance(fecha_raw, datetime):
                    fecha = fecha_raw.strftime("%Y-%m-%d")
                else:
                    fecha = _parse_fecha(str(fecha_raw))

                factura = {
                    "numero_factura": f"XLS-{hash_c[:10]}",
                    "family_code": codigo,
                    "fecha": fecha,
                    "concepto": sheet,
                    "base_imponible": None,
                    "iva": None,
                    "total": total_val,
                    "fecha_vencimiento": None,
                    "estado": "pendiente",
                    "hash_contenido": hash_c,
                    "origen": "excel",
                }

                if _insert_factura(conn, factura):
                    stats["nuevas"] += 1
                else:
                    stats["duplicadas"] += 1
            except Exception as e:
                log.error("Error en fila %d de '%s': %s", i, sheet, e)
                stats["errores"] += 1

    wb.close()
    conn.commit()
    log.info("Excel %s: %d nuevas, %d duplicadas, %d errores",
             filepath.name, stats["nuevas"], stats["duplicadas"], stats["errores"])
    return stats


def procesar_fichero(filepath: Path, conn: sqlite3.Connection) -> dict | None:
    """Despacha el procesamiento seg\u00fan la extensi\u00f3n del fichero."""
    ext = filepath.suffix.lower()
    if ext == ".csv":
        return procesar_csv(filepath, conn)
    elif ext == ".xml":
        return procesar_sepa_xml(filepath, conn)
    elif ext in (".xlsx", ".xls"):
        return procesar_excel(filepath, conn)
    else:
        log.warning("Formato no soportado: %s", filepath.name)
        return None


def procesar_directorio(input_dir: Path, conn: sqlite3.Connection) -> dict:
    """Procesa todos los ficheros compatibles de un directorio."""
    total = {"nuevas": 0, "duplicadas": 0, "errores": 0}
    extensiones = {".csv", ".xml", ".xlsx", ".xls"}

    ficheros = sorted(f for f in input_dir.iterdir()
                      if f.is_file() and f.suffix.lower() in extensiones)

    if not ficheros:
        log.info("No se encontraron ficheros para procesar en %s", input_dir)
        return total

    log.info("Encontrados %d ficheros para procesar", len(ficheros))
    for f in ficheros:
        resultado = procesar_fichero(f, conn)
        if resultado:
            for k in total:
                total[k] += resultado[k]

    log.info("Total: %d nuevas, %d duplicadas, %d errores",
             total["nuevas"], total["duplicadas"], total["errores"])
    return total


def modo_watch(input_dir: Path, conn: sqlite3.Connection):
    """Vigila un directorio y procesa ficheros nuevos autom\u00e1ticamente."""
    try:
        from watchdog.observers import Observer
        from watchdog.events import FileSystemEventHandler
    except ImportError:
        log.error("watchdog no instalado. Ejecuta: pip install watchdog")
        sys.exit(1)

    class Handler(FileSystemEventHandler):
        def on_created(self, event):
            if event.is_directory:
                return
            filepath = Path(event.src_path)
            if filepath.suffix.lower() in {".csv", ".xml", ".xlsx", ".xls"}:
                time.sleep(1)  # Esperar a que el fichero se escriba completamente
                log.info("Nuevo fichero detectado: %s", filepath.name)
                procesar_fichero(filepath, conn)

    observer = Observer()
    observer.schedule(Handler(), str(input_dir), recursive=False)
    observer.start()
    log.info("Vigilando directorio: %s (Ctrl+C para salir)", input_dir)
    try:
        while True:
            time.sleep(1)
    except KeyboardInterrupt:
        observer.stop()
    observer.join()


def main():
    parser = argparse.ArgumentParser(
        description="Procesador de ficheros Sage 50 (CSV, SEPA XML, Excel)"
    )
    parser.add_argument(
        "--input-dir",
        type=Path,
        help="Directorio con ficheros a procesar (modo batch)",
    )
    parser.add_argument(
        "--watch",
        type=Path,
        help="Directorio a vigilar para procesar ficheros nuevos autom\u00e1ticamente",
    )
    parser.add_argument(
        "--db",
        type=Path,
        default=DB_PATH,
        help="Ruta a la base de datos SQLite (por defecto: db/retamar.db)",
    )
    args = parser.parse_args()

    if not args.input_dir and not args.watch:
        parser.error("Especifica --input-dir o --watch")

    conn = init_db(args.db)

    if args.watch:
        modo_watch(args.watch, conn)
    else:
        procesar_directorio(args.input_dir, conn)

    conn.close()


if __name__ == "__main__":
    main()
